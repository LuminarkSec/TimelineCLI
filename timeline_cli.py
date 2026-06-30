#!/usr/bin/env python3
"""
Timeline CLI

Standalone command-line timeline renderer.

Reads timestamped rows from an Excel workbook or CSV file and writes a vertical
timeline as PNG or SVG. It can also export parsed timeline events as JSON.

"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import re
import sys
import textwrap
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.colors import is_color_like


DEFAULT_SHEET = "Timeline"
DEFAULT_TIMESTAMP_COL = "Timestamp_UTC_0"
DEFAULT_ACTIVITY_COL = "Activity"
DEFAULT_TACTIC_COL = "MITRE Tactic"
DEFAULT_VISUALIZE_COL = "Visualize"
DEFAULT_VISUALIZE_VALUE = "yes"

DEFAULT_WIDTH = 14.0
DEFAULT_MIN_HEIGHT = 8.0
DEFAULT_HEIGHT_PER_EVENT = 1.0
DEFAULT_DPI = 150

DEFAULT_BACKGROUND_COLOUR = "white"
DEFAULT_LINE_COLOUR = "black"
DEFAULT_MARKER_COLOUR = "blue"
DEFAULT_TIMESTAMP_COLOUR = "red"
DEFAULT_BAR_COLOUR = "darkorange"
DEFAULT_BAR_TEXT_COLOUR = "white"
DEFAULT_TEXT_COLOUR = "black"
DEFAULT_TITLE_COLOUR = "black"

LABEL_MAX_CHARS = 40
DESC_DISPLAY_MAX = 300

SUPPORTED_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
SUPPORTED_IMAGE_FORMATS = {"png", "svg"}

logger = logging.getLogger("timeline_cli")


class CliError(Exception):
    """User-facing CLI error."""


@dataclass(frozen=True)
class TimelineEvent:
    timestamp: datetime
    activity: str
    tactic: str
    row_number: int


@dataclass(frozen=True)
class TimelineColours:
    background: str = DEFAULT_BACKGROUND_COLOUR
    line: str = DEFAULT_LINE_COLOUR
    marker: str = DEFAULT_MARKER_COLOUR
    timestamp: str = DEFAULT_TIMESTAMP_COLOUR
    bar: str = DEFAULT_BAR_COLOUR
    bar_text: str = DEFAULT_BAR_TEXT_COLOUR
    text: str = DEFAULT_TEXT_COLOUR
    title: str = DEFAULT_TITLE_COLOUR


@dataclass
class BuildStats:
    total_rows_seen: int = 0
    included_events: int = 0
    skipped_visualize_filter: int = 0
    skipped_missing_required: int = 0
    skipped_bad_timestamp: int = 0
    skipped_other_error: int = 0


def strip_fractional_seconds_text(value: str) -> str:
    """
    Strip fractional seconds from common workbook timestamp strings.

    Examples:
      2026-06-23T00:31:53.2971736Z -> 2026-06-23T00:31:53Z
      2026-06-23 00:31:53.297173  -> 2026-06-23 00:31:53
      2026-06-23 00:31:53:297173  -> 2026-06-23 00:31:53
    """
    text = str(value).strip()

    text = re.sub(
        r"(\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}:\d{2})\.\d+",
        r"\1",
        text,
    )

    text = re.sub(
        r"(\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}:\d{2}):\d+",
        r"\1",
        text,
    )

    return text


def parse_workbook_datetime(value: Any) -> datetime:
    """
    Parse workbook timestamp values.

    Supports:
      - Excel/openpyxl datetime cells
      - 2026-06-23 00:31:53
      - 2026-06-23 00:31:53.297173
      - 2026-06-23 00:31:53:297173
      - 2026-06-23
      - 2026-06-23T00:31:53.2971736Z
      - 2026-06-23T00:31:53Z
      - 2026-06-23T01:31:53+01:00

    Returns naive UTC datetime values with microseconds stripped.
    """
    if isinstance(value, datetime):
        dt = value
    else:
        raw = strip_fractional_seconds_text(str(value).strip())

        if not raw:
            raise ValueError("empty timestamp")

        iso_value = raw.replace(" ", "T")

        if iso_value.endswith(("Z", "z")):
            iso_value = iso_value[:-1] + "+00:00"

        try:
            dt = datetime.fromisoformat(iso_value)
        except ValueError:
            fallback_formats = (
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d",
            )

            for fmt in fallback_formats:
                try:
                    dt = datetime.strptime(raw, fmt)
                    break
                except ValueError:
                    continue
            else:
                raise ValueError(f"unsupported timestamp format: {value!r}")

    if dt.tzinfo is not None:
        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)

    return dt.replace(microsecond=0)


def format_timestamp(value: datetime, iso: bool = False) -> str:
    return value.isoformat() if iso else value.strftime("%Y-%m-%d %H:%M:%S")


def clean_text(value: Any, max_chars: Optional[int] = None) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\t", " ").replace("\n", " ").replace("\r", " ")
    text = " ".join(text.split())

    if max_chars is not None and len(text) > max_chars:
        return text[: max_chars - 3] + "..."

    return text


def normalise_header(value: Any, index: int) -> str:
    if value is None or str(value).strip() == "":
        return f"Column {index + 1}"

    return str(value).strip()


def validate_colour(value: str, option_name: str) -> None:
    if not is_color_like(value):
        raise CliError(
            f"{option_name} value {value!r} is not a valid Matplotlib colour. "
            "Use a named colour such as 'red', 'black', 'darkorange', "
            "or a hex value such as '#ff0000'."
        )


def load_workbook_for_metadata(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise CliError("Excel input requires openpyxl. Install it with: pip install openpyxl") from exc

    try:
        return load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise CliError(f"Could not read workbook {path}: {exc}") from exc


def read_xlsx_headers(path: Path, sheet_name: str) -> List[str]:
    workbook = load_workbook_for_metadata(path)

    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise CliError(f"Sheet {sheet_name!r} not found. Available sheets: {available}")

    sheet = workbook[sheet_name]

    try:
        header_row = next(sheet.iter_rows(values_only=True))
    except StopIteration as exc:
        raise CliError(f"Sheet {sheet_name!r} is empty") from exc

    return [normalise_header(h, i) for i, h in enumerate(header_row)]


def read_csv_headers(path: Path) -> List[str]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)

            if not reader.fieldnames:
                raise CliError(f"CSV file {path} has no header row")

            return [str(h).strip() for h in reader.fieldnames]

    except UnicodeDecodeError as exc:
        raise CliError(f"Could not decode CSV file {path}; expected UTF-8 or UTF-8 with BOM") from exc
    except OSError as exc:
        raise CliError(f"Could not read CSV file {path}: {exc}") from exc


def read_headers(path: Path, sheet_name: str) -> List[str]:
    suffix = path.suffix.lower()

    if suffix in SUPPORTED_EXCEL_SUFFIXES:
        return read_xlsx_headers(path, sheet_name)

    if suffix == ".csv":
        return read_csv_headers(path)

    raise CliError(f"Unsupported input type {suffix!r}. Use .xlsx/.xlsm/.xltx/.xltm or .csv.")


def load_rows_from_xlsx(path: Path, sheet_name: str) -> Iterable[Dict[str, Any]]:
    workbook = load_workbook_for_metadata(path)

    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise CliError(f"Sheet {sheet_name!r} not found. Available sheets: {available}")

    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise CliError(f"Sheet {sheet_name!r} is empty") from exc

    headers = [normalise_header(h, i) for i, h in enumerate(header_row)]

    for excel_row_number, row in enumerate(rows, start=2):
        yield {
            "_row_number": excel_row_number,
            **{
                headers[i]: row[i] if i < len(row) else None
                for i in range(len(headers))
            },
        }


def load_rows_from_csv(path: Path) -> Iterable[Dict[str, Any]]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)

            if not reader.fieldnames:
                raise CliError(f"CSV file {path} has no header row")

            for csv_row_number, row in enumerate(reader, start=2):
                yield {"_row_number": csv_row_number, **row}

    except UnicodeDecodeError as exc:
        raise CliError(f"Could not decode CSV file {path}; expected UTF-8 or UTF-8 with BOM") from exc
    except OSError as exc:
        raise CliError(f"Could not read CSV file {path}: {exc}") from exc


def load_rows(path: Path, sheet_name: str) -> Iterable[Dict[str, Any]]:
    suffix = path.suffix.lower()

    if suffix in SUPPORTED_EXCEL_SUFFIXES:
        yield from load_rows_from_xlsx(path, sheet_name)
        return

    if suffix == ".csv":
        yield from load_rows_from_csv(path)
        return

    raise CliError(f"Unsupported input type {suffix!r}. Use .xlsx/.xlsm/.xltx/.xltm or .csv.")


def require_columns(headers: Sequence[str], required_columns: Sequence[str]) -> None:
    missing = [col for col in required_columns if col and col not in headers]

    if not missing:
        return

    available = "\n".join(f"  - {h}" for h in headers)
    missing_list = ", ".join(repr(col) for col in missing)

    raise CliError(
        f"Missing required column(s): {missing_list}\n\n"
        f"Available columns:\n{available}"
    )


def parse_visualize_values(raw_value: str) -> set[str]:
    values = {
        value.strip().lower()
        for value in raw_value.split(",")
        if value.strip()
    }

    if not values:
        raise CliError("--visualize-value must contain at least one non-empty value")

    return values


def build_events(
    rows: Iterable[Dict[str, Any]],
    timestamp_col: str,
    activity_col: str,
    tactic_col: str,
    visualize_col: Optional[str],
    visualize_values: set[str],
    use_visualize_filter: bool,
    max_rows: int,
) -> Tuple[List[TimelineEvent], BuildStats]:
    events: List[TimelineEvent] = []
    stats = BuildStats()

    for scanned_index, row in enumerate(rows, start=1):
        if scanned_index > max_rows:
            break

        stats.total_rows_seen += 1
        row_number = int(row.get("_row_number") or scanned_index + 1)

        try:
            if use_visualize_filter and visualize_col:
                raw_visualize = row.get(visualize_col)
                if str(raw_visualize).strip().lower() not in visualize_values:
                    stats.skipped_visualize_filter += 1
                    continue

            timestamp_value = row.get(timestamp_col)
            activity_value = row.get(activity_col)
            tactic_value = row.get(tactic_col)

            if not timestamp_value or not activity_value or not tactic_value:
                stats.skipped_missing_required += 1
                continue

            try:
                timestamp = parse_workbook_datetime(timestamp_value)
            except ValueError as exc:
                stats.skipped_bad_timestamp += 1
                logger.warning(
                    "Skipping row %s: invalid timestamp %r (%s)",
                    row_number,
                    timestamp_value,
                    exc,
                )
                continue

            events.append(
                TimelineEvent(
                    timestamp=timestamp,
                    activity=clean_text(activity_value),
                    tactic=clean_text(tactic_value),
                    row_number=row_number,
                )
            )

        except Exception as exc:
            stats.skipped_other_error += 1
            logger.warning("Skipping row %s: %s", row_number, exc)

    events.sort(key=lambda event: event.timestamp)
    stats.included_events = len(events)

    return events, stats


def events_to_json_data(events: Sequence[TimelineEvent]) -> List[Dict[str, Any]]:
    return [
        {
            "timestamp": format_timestamp(event.timestamp, iso=True),
            "timestamp_display": format_timestamp(event.timestamp),
            "activity": event.activity,
            "mitre_tactic": event.tactic,
            "source_row": event.row_number,
        }
        for event in events
    ]


def write_json(events: Sequence[TimelineEvent], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    data = events_to_json_data(events)
    output_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def infer_output_format(output_path: Optional[Path], explicit_format: Optional[str]) -> str:
    if explicit_format:
        fmt = explicit_format.lower()
    elif output_path and output_path.suffix:
        fmt = output_path.suffix.lower().lstrip(".")
    else:
        fmt = "png"

    if fmt not in SUPPORTED_IMAGE_FORMATS:
        allowed = ", ".join(sorted(SUPPORTED_IMAGE_FORMATS))
        raise CliError(f"Unsupported output format {fmt!r}. Supported formats: {allowed}")

    return fmt


def paged_output_path(output_path: Path, page_number: int, total_pages: int, output_format: str) -> Path:
    suffix = f".{output_format}"

    if total_pages <= 1:
        if output_path.suffix.lower().lstrip(".") == output_format:
            return output_path
        return output_path.with_suffix(suffix)

    stem = output_path.stem if output_path.suffix else output_path.name
    parent = output_path.parent

    return parent / f"{stem}_{page_number:03d}{suffix}"


def chunk_events(events: Sequence[TimelineEvent], page_size: Optional[int]) -> List[Sequence[TimelineEvent]]:
    if not page_size or page_size <= 0 or page_size >= len(events):
        return [events]

    return [
        events[start : start + page_size]
        for start in range(0, len(events), page_size)
    ]


def draw_timeline(
    events: Sequence[TimelineEvent],
    output_path: Path,
    title: str,
    width: float,
    min_height: float,
    height_per_event: float,
    dpi: int,
    output_format: str,
    colours: TimelineColours,
    page_number: int = 1,
    total_pages: int = 1,
) -> None:
    if not events:
        raise CliError("No timeline events were supplied to renderer")

    margin = 50
    line_x = 400
    marker_size = 7
    text_offset = 20
    timestamp_width = 2
    tactic_width = 325
    tactic_height = 20
    desc_width = 550
    min_spacing = 100

    # Estimate wrapped activity height so long activities get more vertical room,
    activity_heights = []

    for event in events:
        description = clean_text(event.activity, DESC_DISPLAY_MAX)
        wrapped_lines = textwrap.wrap(description, width=85) or [""]
        desc_height = len(wrapped_lines) * 16 + 10

        # Base row height:
        # 30px for marker/timestamp/tactic row + wrapped description height.
        row_height = 30 + desc_height
        activity_heights.append(max(row_height, 60))

    y_positions = [margin]

    for index in range(1, len(events)):
        previous_y = y_positions[-1]
        previous_height = activity_heights[index - 1]
        y_positions.append(previous_y + previous_height + 20)

    canvas_height = max(
        1000,
        int(y_positions[-1] + activity_heights[-1] + margin),
    )
    canvas_width = 1100

    # Convert the pixel-like canvas into a Matplotlib figure.
    fig_width = width
    fig_height = max(min_height, canvas_height / 100)
    fig, ax = plt.subplots(figsize=(fig_width, fig_height))

    fig.patch.set_facecolor(colours.background)
    ax.set_facecolor(colours.background)

    # Draw the vertical timeline only over actual content.
    line_start = y_positions[0]
    line_end = y_positions[-1] + activity_heights[-1]

    ax.plot(
        [line_x, line_x],
        [line_start, line_end],
        color=colours.line,
        linewidth=2,
    )

    for event, y_pos in zip(events, y_positions):
        # Marker
        ax.plot(
            line_x,
            y_pos,
            "o",
            color=colours.marker,
            markersize=marker_size,
        )

        # Timestamp, left of the line.
        ax.text(
            line_x - timestamp_width - text_offset,
            y_pos,
            format_timestamp(event.timestamp),
            fontsize=10,
            ha="right",
            va="center",
            color=colours.timestamp,
            weight="bold",
            family="sans-serif",
        )

        # Fixed-width tactic bar.
        tactic = clean_text(event.tactic, LABEL_MAX_CHARS)
        tactic_x = line_x + text_offset
        tactic_y = y_pos - 10

        ax.add_patch(
            plt.Rectangle(
                (tactic_x, tactic_y),
                tactic_width,
                tactic_height,
                facecolor=colours.bar,
                edgecolor="none",
            )
        )

        # Tactic text is left-aligned inside the fixed bar
        ax.text(
            tactic_x + 5,
            y_pos,
            tactic,
            fontsize=10,
            ha="left",
            va="center",
            color=colours.bar_text,
            weight="bold",
            family="sans-serif",
        )

        # Activity underneath the tactic bar.
        description = clean_text(event.activity, DESC_DISPLAY_MAX)
        wrapped_description = "\n".join(textwrap.wrap(description, width=85)) if description else ""

        activity_x = line_x + text_offset
        activity_y = y_pos + 15

        ax.text(
            activity_x,
            activity_y,
            wrapped_description,
            fontsize=10,
            ha="left",
            va="top",
            color=colours.text,
            family="sans-serif",
        )

    page_suffix = f" — page {page_number}/{total_pages}" if total_pages > 1 else ""

    ax.set_xlim(0, canvas_width)

    # Invert Y so positive values go down the page.
    ax.set_ylim(canvas_height, 0)

    ax.axis("off")
    ax.set_title(
        f"{title}{page_suffix}",
        fontsize=14,
        weight="bold",
        pad=20,
        color=colours.title,
        family="sans-serif",
    )

    plt.tight_layout()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    plt.savefig(
        output_path,
        format=output_format,
        dpi=dpi,
        bbox_inches="tight",
        facecolor=colours.background,
    )
    plt.close(fig)


def render_outputs(
    events: Sequence[TimelineEvent],
    output_path: Path,
    output_format: str,
    page_size: Optional[int],
    title: str,
    width: float,
    min_height: float,
    height_per_event: float,
    dpi: int,
    colours: TimelineColours,
) -> List[Path]:
    pages = chunk_events(events, page_size)
    total_pages = len(pages)
    written_paths: List[Path] = []

    for idx, page_events in enumerate(pages, start=1):
        page_output = paged_output_path(output_path, idx, total_pages, output_format)

        draw_timeline(
            events=page_events,
            output_path=page_output,
            title=title,
            width=width,
            min_height=min_height,
            height_per_event=height_per_event,
            dpi=dpi,
            output_format=output_format,
            colours=colours,
            page_number=idx,
            total_pages=total_pages,
        )

        written_paths.append(page_output)

    return written_paths


def print_stats(stats: BuildStats) -> None:
    print(
        "Rows: "
        f"seen={stats.total_rows_seen}, "
        f"included={stats.included_events}, "
        f"skipped_visualize={stats.skipped_visualize_filter}, "
        f"skipped_missing_required={stats.skipped_missing_required}, "
        f"skipped_bad_timestamp={stats.skipped_bad_timestamp}, "
        f"skipped_other={stats.skipped_other_error}"
    )


def print_list(items: Sequence[str], heading: str) -> None:
    print(heading)

    for item in items:
        print(f"  - {item}")


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create a standalone timeline PNG/SVG from an Excel workbook or CSV file.",
    )

    parser.add_argument(
        "input",
        type=Path,
        help="Input .xlsx/.xlsm workbook or .csv file",
    )

    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Output image path. Required unless using --json-only, --list-sheets, or --list-columns.",
    )

    parser.add_argument(
        "--format",
        choices=sorted(SUPPORTED_IMAGE_FORMATS),
        help="Output image format. Defaults to output extension, then png.",
    )

    parser.add_argument(
        "--json",
        type=Path,
        dest="json_output",
        help="Also write parsed timeline events to this JSON file",
    )

    parser.add_argument(
        "--json-only",
        action="store_true",
        help="Only write JSON; do not render an image. Requires --json.",
    )

    parser.add_argument(
        "--page-size",
        type=int,
        help="Split output into multiple files with this many events per image, e.g. --page-size 50",
    )

    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET,
        help=f"Workbook sheet name, default: {DEFAULT_SHEET}",
    )

    parser.add_argument(
        "--timestamp-col",
        default=DEFAULT_TIMESTAMP_COL,
        help=f"Timestamp column, default: {DEFAULT_TIMESTAMP_COL}",
    )

    parser.add_argument(
        "--activity-col",
        default=DEFAULT_ACTIVITY_COL,
        help=f"Activity column, default: {DEFAULT_ACTIVITY_COL}",
    )

    parser.add_argument(
        "--tactic-col",
        default=DEFAULT_TACTIC_COL,
        help=f"Category/tactic column, default: {DEFAULT_TACTIC_COL}",
    )

    parser.add_argument(
        "--visualize-col",
        default=DEFAULT_VISUALIZE_COL,
        help=f"Filter column, default: {DEFAULT_VISUALIZE_COL}",
    )

    parser.add_argument(
        "--visualize-value",
        default=DEFAULT_VISUALIZE_VALUE,
        help="Accepted filter value(s), comma-separated. Default: yes",
    )

    parser.add_argument(
        "--no-visualize-filter",
        action="store_true",
        help="Include all rows instead of filtering on the filter column",
    )

    parser.add_argument(
        "--list-sheets",
        action="store_true",
        help="List workbook sheet names and exit",
    )

    parser.add_argument(
        "--list-columns",
        action="store_true",
        help="List input columns for the selected sheet/CSV and exit",
    )

    parser.add_argument(
        "--max-rows",
        type=int,
        default=1000,
        help="Maximum input rows to scan, default: 1000",
    )

    parser.add_argument(
        "--title",
        default="Timeline Visualization",
        help="Chart title",
    )

    parser.add_argument(
        "--width",
        type=float,
        default=DEFAULT_WIDTH,
        help=f"Figure width in inches, default: {DEFAULT_WIDTH}",
    )

    parser.add_argument(
        "--min-height",
        type=float,
        default=DEFAULT_MIN_HEIGHT,
        help=f"Minimum figure height in inches, default: {DEFAULT_MIN_HEIGHT}",
    )

    parser.add_argument(
        "--height-per-event",
        type=float,
        default=DEFAULT_HEIGHT_PER_EVENT,
        help=f"Height per event in inches, default: {DEFAULT_HEIGHT_PER_EVENT}",
    )

    parser.add_argument(
        "--dpi",
        type=int,
        default=DEFAULT_DPI,
        help=f"Output DPI for raster formats, default: {DEFAULT_DPI}",
    )

    parser.add_argument(
        "--background-colour",
        "--background-color",
        default=DEFAULT_BACKGROUND_COLOUR,
        help=f"Timeline background colour, default: {DEFAULT_BACKGROUND_COLOUR}",
    )

    parser.add_argument(
        "--line-colour",
        "--line-color",
        default=DEFAULT_LINE_COLOUR,
        help=f"Vertical timeline line colour, default: {DEFAULT_LINE_COLOUR}",
    )

    parser.add_argument(
        "--marker-colour",
        "--marker-color",
        default=DEFAULT_MARKER_COLOUR,
        help=f"Timeline marker colour, default: {DEFAULT_MARKER_COLOUR}",
    )

    parser.add_argument(
        "--timestamp-colour",
        "--timestamp-color",
        default=DEFAULT_TIMESTAMP_COLOUR,
        help=f"Timestamp text colour, default: {DEFAULT_TIMESTAMP_COLOUR}",
    )

    parser.add_argument(
        "--bar-colour",
        "--bar-color",
        default=DEFAULT_BAR_COLOUR,
        help=f"Category/tactic bar colour, default: {DEFAULT_BAR_COLOUR}",
    )

    parser.add_argument(
        "--bar-text-colour",
        "--bar-text-color",
        default=DEFAULT_BAR_TEXT_COLOUR,
        help=f"Category/tactic bar text colour, default: {DEFAULT_BAR_TEXT_COLOUR}",
    )

    parser.add_argument(
        "--text-colour",
        "--text-color",
        default=DEFAULT_TEXT_COLOUR,
        help=f"Activity text colour, default: {DEFAULT_TEXT_COLOUR}",
    )

    parser.add_argument(
        "--title-colour",
        "--title-color",
        default=DEFAULT_TITLE_COLOUR,
        help=f"Title text colour, default: {DEFAULT_TITLE_COLOUR}",
    )

    parser.add_argument(
        "--stats",
        action="store_true",
        help="Print row processing statistics",
    )

    parser.add_argument(
        "-v",
        "--verbose",
        action="store_true",
        help="Show row skip warnings",
    )

    return parser.parse_args(argv)


def validate_args(args: argparse.Namespace) -> None:
    if not args.input.is_file():
        raise CliError(f"Input file not found: {args.input}")

    if args.json_only and not args.json_output:
        raise CliError("--json-only requires --json")

    if not args.json_only and not args.output and not args.list_sheets and not args.list_columns:
        raise CliError("--output is required unless using --json-only, --list-sheets, or --list-columns")

    if args.page_size is not None and args.page_size <= 0:
        raise CliError("--page-size must be greater than zero")

    if args.max_rows <= 0:
        raise CliError("--max-rows must be greater than zero")

    if args.width <= 0:
        raise CliError("--width must be greater than zero")

    if args.min_height <= 0:
        raise CliError("--min-height must be greater than zero")

    if args.height_per_event <= 0:
        raise CliError("--height-per-event must be greater than zero")

    if args.dpi <= 0:
        raise CliError("--dpi must be greater than zero")

    colour_options = {
        "--background-colour": args.background_colour,
        "--line-colour": args.line_colour,
        "--marker-colour": args.marker_colour,
        "--timestamp-colour": args.timestamp_colour,
        "--bar-colour": args.bar_colour,
        "--bar-text-colour": args.bar_text_colour,
        "--text-colour": args.text_colour,
        "--title-colour": args.title_colour,
    }

    for option_name, colour_value in colour_options.items():
        validate_colour(colour_value, option_name)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)

    logging.basicConfig(
        level=logging.INFO if args.verbose else logging.WARNING,
        format="%(levelname)s: %(message)s",
    )

    try:
        validate_args(args)

        if args.list_sheets:
            if args.input.suffix.lower() not in SUPPORTED_EXCEL_SUFFIXES:
                raise CliError("--list-sheets is only available for Excel input")

            workbook = load_workbook_for_metadata(args.input)
            print_list(workbook.sheetnames, "Sheets:")
            return 0

        if args.list_columns:
            headers = read_headers(args.input, args.sheet)
            print_list(headers, "Columns:")
            return 0

        headers = read_headers(args.input, args.sheet)

        required_columns = [
            args.timestamp_col,
            args.activity_col,
            args.tactic_col,
        ]

        if not args.no_visualize_filter and args.visualize_col:
            required_columns.append(args.visualize_col)

        require_columns(headers, required_columns)

        visualize_values = parse_visualize_values(args.visualize_value)

        colours = TimelineColours(
            background=args.background_colour,
            line=args.line_colour,
            marker=args.marker_colour,
            timestamp=args.timestamp_colour,
            bar=args.bar_colour,
            bar_text=args.bar_text_colour,
            text=args.text_colour,
            title=args.title_colour,
        )

        rows = load_rows(args.input, args.sheet)

        events, stats = build_events(
            rows=rows,
            timestamp_col=args.timestamp_col,
            activity_col=args.activity_col,
            tactic_col=args.tactic_col,
            visualize_col=args.visualize_col,
            visualize_values=visualize_values,
            use_visualize_filter=not args.no_visualize_filter,
            max_rows=args.max_rows,
        )

        if args.stats:
            print_stats(stats)

        if not events:
            raise CliError(
                "No timeline events found.\n\n"
                "Checks:\n"
                f"  - Sheet/CSV has column {args.timestamp_col!r}\n"
                f"  - Sheet/CSV has column {args.activity_col!r}\n"
                f"  - Sheet/CSV has column {args.tactic_col!r}\n"
                f"  - Rows contain non-empty values in those columns\n"
                f"  - If filtering is enabled, {args.visualize_col!r} matches one of: "
                f"{', '.join(sorted(visualize_values))}\n\n"
                "Tip: rerun with --stats -v, or use --no-visualize-filter."
            )

        if args.json_output:
            write_json(events, args.json_output)
            print(f"Wrote JSON {args.json_output} with {len(events)} event(s)")

        if not args.json_only:
            assert args.output is not None

            output_format = infer_output_format(args.output, args.format)

            written_paths = render_outputs(
                events=events,
                output_path=args.output,
                output_format=output_format,
                page_size=args.page_size,
                title=args.title,
                width=args.width,
                min_height=args.min_height,
                height_per_event=args.height_per_event,
                dpi=args.dpi,
                colours=colours,
            )

            if len(written_paths) == 1:
                print(f"Wrote {written_paths[0]} with {len(events)} event(s)")
            else:
                print(f"Wrote {len(written_paths)} {output_format.upper()} page(s):")
                for path in written_paths:
                    print(f"  - {path}")

        return 0

    except CliError as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 2
    except KeyboardInterrupt:
        print("Interrupted", file=sys.stderr)
        return 130


if __name__ == "__main__":
    raise SystemExit(main())
